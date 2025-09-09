import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os
import datetime


# ============
wb_path = r"dataset_90/1_110-2_answer.xlsx"
workbook = openpyxl.load_workbook(wb_path)

# ============
sheet = workbook["Sheet1"]
sum_column = sheet.max_column + 1

# ============
for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
    row_values = []
    for value in row:
        if value is not None and isinstance(value, (int, float)):
            row_values.append(value)
    if row_values:
        sum_of_values = sum(row_values)
        sheet.cell(row=row_num + 1, column=sum_column).value = sum_of_values

# ============
print(sheet.max_column)

# ============
workbook.save(r"results1/SENZA_preprocessing/1_110-2_answer/workbook_new.xlsx")