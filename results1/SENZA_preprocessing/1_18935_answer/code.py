import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os
import datetime


# ============
wb_path = r"dataset_90/1_18935_answer.xlsx"
workbook = openpyxl.load_workbook(wb_path)

# ============
sheet_names = workbook.sheetnames
print(sheet_names)

# ============
ws = workbook["Sheet1"]
max_row = ws.max_row
print(max_row)

# ============
ws = workbook["Sheet1"]
header_row = [cell.value for cell in ws[1]]
print(header_row)

# ============
ws = workbook["Sheet1"]
category_value_column_index = 7

# ============
ws = workbook["Sheet1"]
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=category_value_column_index, max_col=category_value_column_index):
    for cell in row:
        cell.value = None

# ============
workbook.save(r"results1/SENZA_preprocessing/1_18935_answer/workbook_new.xlsx")

# ============
workbook.save(r"results1/SENZA_preprocessing/1_18935_answer/workbook_new.xlsx")