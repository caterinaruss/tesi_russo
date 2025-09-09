# -*- coding: utf-8 -*-
import openpyxl
import numpy as np
from sklearn.cluster import DBSCAN

def detect_tables_in_sheet(sheet, eps_range=(1.0, 2.0), min_samples_range=(2, 5)):
    """
    rilevamento di tabelle in un singolo foglio Excel.
    Utilizza contenuto, bordi, grassetto e colore di sfondo per l'analisi.
    Ottimizza i parametri DBSCAN per trovare la configurazione migliore.
    
    Returns:
        tuple: (parametri_ottimali, tutti_risultati)
        parametri_ottimali = (eps, min_samples, bboxes)
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    if max_row == 0 or max_col == 0:
        return None, []

    # Costruisci una matrice di punteggi basata su contenuto, bordi e stile
    score_matrix = np.zeros((max_row, max_col), dtype=int)
    for r_idx in range(1, max_row + 1):
        for c_idx in range(1, max_col + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            score = 0
            
            # Punteggio per il contenuto
            if cell.value is not None and str(cell.value).strip() != "":
                score += 1
            
            # Punteggio per i bordi
            if cell.border.left.style: score += 2
            if cell.border.right.style: score += 2
            if cell.border.top.style: score += 2
            if cell.border.bottom.style: score += 2
            
            # Punteggio per lo stile del font (es. grassetto)
            if cell.font and cell.font.bold:
                score += 3  # Punteggio alto per il grassetto, forte indicatore di intestazione

            # Punteggio per il colore di sfondo (fill)
            if cell.fill and cell.fill.fill_type and cell.fill.fill_type != 'none':
                score += 1 # Punteggio più basso, a volte è solo decorativo
            
            score_matrix[r_idx-1, c_idx-1] = score

    # Estrai coordinate delle celle con un punteggio > 0 ("celle interessanti")
    interesting_cells = []
    for i in range(max_row):
        for j in range(max_col):
            if score_matrix[i, j] > 0:
                interesting_cells.append([i, j])
    
    if not interesting_cells:
        return None, []
    
    X = np.array(interesting_cells)
    
    # Test sistematico di parametri DBSCAN
    best_params = None
    best_score = float('inf')
    all_results = []
    
    eps_values = np.arange(eps_range[0], eps_range[1] + 0.1, 0.1)
    min_samples_values = range(min_samples_range[0], min_samples_range[1] + 1)
    
    for eps in eps_values:
        for min_samp in min_samples_values:
            clustering = DBSCAN(eps=eps, min_samples=min_samp).fit(X)
            labels = clustering.labels_
            
            # Calcola bounding box delle tabelle
            table_bboxes = []
            for label in set(labels):
                if label == -1:
                    continue  # ignora outliers
                points = X[labels == label]
                min_row, min_col = points.min(axis=0)
                max_row, max_col = points.max(axis=0)
                # Bbox in formato [r0, c0, r1, c1] con indici base 1
                table_bboxes.append([int(min_row+1), int(min_col+1), int(max_row+1), int(max_col+1)])
            
            num_tables = len(table_bboxes)
            num_outliers = np.sum(labels == -1)
            
            # Scoring: premia un numero ragionevole di tabelle e penalizza outliers
            # Un valore negativo alto per num_tables incentiva a trovare cluster
            score = num_outliers - num_tables * 5 
            
            all_results.append({
                'eps': eps, 
                'min_samples': min_samp, 
                'num_tables': num_tables, 
                'num_outliers': num_outliers,
                'score': score,
                'bboxes': table_bboxes
            })
            
            if score < best_score and num_tables > 0: # Assicurati di trovare almeno una tabella
                best_score = score
                best_params = (eps, min_samp, table_bboxes)
    
    return best_params, all_results

def analyze_workbook(file_path, eps_range=(1.0, 2.0), min_samples_range=(2, 5), verbose=True):
    """
    Analizza un intero file Excel (tutti i fogli) e rileva le tabelle.

    Returns:
        dict con risultati dettagliati per l'intero workbook
    """
    import json
    
    if verbose:
        print(f"Analizzando workbook: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        if verbose:
            print(f"Errore nell'aprire il file {file_path}: {e}")
        return None

    workbook_results = {
        'file_path': file_path,
        'sheets': []
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if verbose:
            print(f"--- Analizzando foglio: {sheet_name} ---")

        best_params, all_results = detect_tables_in_sheet(
            sheet, eps_range, min_samples_range
        )

        sheet_summary = {'sheet_name': sheet_name}
        if not best_params:
            if verbose:
                print("Nessuna tabella trovata.")
            sheet_summary.update({'num_tables': 0, 'tables': []})
        else:
            eps, min_samp, bboxes = best_params
            if verbose:
                print(f"Tabelle trovate: {len(bboxes)}")
                for i, bbox in enumerate(bboxes):
                    print(f"  - Tabella {i+1}: Bbox = {bbox}")
            
            tables_data = []
            for idx, bbox in enumerate(bboxes):
                tables_data.append({'table_id': idx + 1, 'bbox': bbox})
            
            sheet_summary.update({
                'num_tables': len(bboxes),
                'parameters': {'eps': eps, 'min_samples': min_samp},
                'tables': tables_data
            })
        
        workbook_results['sheets'].append(sheet_summary)

    return workbook_results
